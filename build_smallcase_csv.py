"""
build_smallcase_csv.py
----------------------
Emit the smallcase upload CSV for the current All-Indices portfolio, in the
same schema as the smallcase template:

    NSE Ticker,Weight,Segment (optional),Rationale (optional)

Holdings come from the live site payload (d:/SQE-host/data.js ->
DASHBOARD_DATA.total759.current_portfolio), so the file always reflects the
month the dashboard is showing. The rationale carries the two numbers the
strategy actually selects on: beta and ERB (excess return to beta).

    python build_smallcase_csv.py
    python build_smallcase_csv.py --out "C:/Users/PC2546/Downloads/x.csv"
"""
import argparse
import csv
import json
import math
import os

DATA_JS = os.environ.get('SQE_DATA_JS', r'd:/SQE-host/data.js')
UNIVERSE = os.environ.get('SQE_UNIVERSE', 'total759')
DP = 2   # decimals on the Weight column, which is a percentage: 10 = 10%

# Company name as registered against the NSE ticker, plus the smallcase-style
# segment. Names verified against the ticker itself, not the local CSV headers
# (those are stale for several of these).
META = {
    'ATHERENERG': ('Ather Energy', 'Two Wheelers'),
    'PIRAMALFIN': ('Piramal Finance', 'NBFC'),
    'BELRISE':    ('Belrise Industries', 'Auto Components'),
    'CORONA':     ('Corona Remedies', 'Pharmaceuticals'),
    'CANHLIFE':   ('Canara HSBC Life Insurance', 'Life Insurance'),
    'RUBICON':    ('Rubicon Research', 'Pharmaceuticals'),
    'PARKHOSPS':  ('Park Medi World', 'Hospitals & Diagnostic Centres'),
    'CPPLUS':     ('Aditya Infotech (CP PLUS)', 'Electronic Equipments'),
    'LENSKART':   ('Lenskart Solutions', 'Retail - Speciality'),
    'EMMVEE':     ('Emmvee Photovoltaic Power', 'Renewable Energy Equipment'),
    'CUPID':      ('Cupid', 'Household & Personal Products'),
    'TVSMOTOR':   ('TVS Motor Company', 'Two Wheelers'),
    'CHOICEIN':   ('Choice International', 'Stockbroking & Allied'),
    'ATLANTAELE': ('Atlanta Electricals', 'Heavy Electrical Equipments'),
    'SKYGOLD':    ('Sky Gold and Diamonds', 'Gems & Jewellery'),
    'AETHER':     ('Aether Industries', 'Specialty Chemicals'),
    'AVL':        ('Aditya Vision', 'Retail - Speciality'),
    'SHAILY':     ('Shaily Engineering Plastics', 'Plastic Products'),
    'JSLL':       ('Jeena Sikho Lifecare', 'Hospitals & Diagnostic Centres'),
    'GOKULAGRO':  ('Gokul Agro Resources', 'Edible Oils'),
    'SMLMAH':     ('SML Mahindra', 'Commercial Vehicles'),
    'PGIL':       ('Pearl Global Industries', 'Apparel & Accessories'),
    'PGEL':       ('PG Electroplast', 'Consumer Electronics'),
    'TARIL':      ('Transformers & Rectifiers India', 'Heavy Electrical Equipments'),
    'APARINDS':   ('APAR Industries', 'Cables'),
    'BLUESTONE':  ('BlueStone Jewellery and Lifestyle', 'Gems & Jewellery'),
    'DIACABS':    ('Diamond Power Infrastructure', 'Cables'),
    'TI':         ('Tilaknagar Industries', 'Breweries & Distilleries'),
    'HCG':        ('HealthCare Global Enterprises', 'Hospitals & Diagnostic Centres'),
    'BSE':        ('BSE', 'Exchanges & Data Platforms'),
    # Entered the book at the September 2026 rebalance.
    'SAILIFE':    ('Sai Life Sciences', 'Pharmaceuticals'),
    'KRN':        ('KRN Heat Exchanger and Refrigeration', 'Industrial Products'),
    'AVALON':     ('Avalon Technologies', 'Electronic Equipments'),
}


def load_portfolio(path, universe):
    s = open(path, encoding='utf-8').read()
    d = json.loads(s[s.index('=') + 1:s.rfind(';')].strip())
    return d[universe]['current_portfolio'], d.get('last_update', '')


def risk_band(beta):
    if beta < 0.35:
        return 'near market-neutral'
    if beta < 0.75:
        return 'low market sensitivity'
    if beta <= 1.15:
        return 'market-like sensitivity'
    return 'high market sensitivity'


def rationale(h, name, segment, qty, cost, exec_w):
    beta, erb = h['beta'], h['erb']
    # prev_qty, not status: the engine marks a first-time buy "Remained" too
    # (SAILIFE / KRN / AVALON entered in Sep'26 that way), so keying off status
    # labelled brand-new positions as carried over. Fall back to status only
    # where prev_qty is absent, for older data.js files.
    prev = h.get('prev_qty')
    fresh = (prev == 0) if prev is not None else h['status'].lower().startswith('add')
    entry = ('Fresh entry this rebalance' if fresh
             else 'Carried over from the previous rebalance')
    return (
        f"{name} ({segment}). Beta {beta:.2f} vs the benchmark "
        f"({risk_band(beta)}); excess return to beta (ERB) {erb:.2f}, i.e. "
        f"{erb:.2f}x of excess return earned per unit of market risk taken. "
        f"Selected by the SQE quant screen, which ranks the All-Indices "
        f"universe on ERB and sizes positions so the highest excess-beta-return "
        f"names carry the most weight. {entry}. Model weight "
        f"{h['weight'] * 100:.2f}%, executed as {qty} share"
        f"{'' if qty == 1 else 's'} at Rs {h['ltp']:,.2f} = Rs {cost:,.0f}, "
        f"which is {exec_w * 100:.2f}% of the capital actually deployed."
    )


def size(port, capital):
    """Whole-share sizing, same rule the dashboard uses (app.js:625):
    qty = max(1, floor(capital * weight / ltp)). The 1-share floor means a
    high-priced, low-weight name can consume far more than its model weight,
    so the deployed total overshoots `capital`."""
    lots = []
    for h in port:
        qty = max(1, math.floor(capital * h['weight'] / h['ltp']))
        lots.append((qty, qty * h['ltp']))
    return lots, sum(c for _, c in lots)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--data', default=DATA_JS)
    ap.add_argument('--universe', default=UNIVERSE)
    ap.add_argument('--capital', type=float, default=100000,
                    help='basket size the whole-share sizing is run at')
    ap.add_argument('--basis', choices=['executed', 'model'], default='executed',
                    help='executed = share of capital actually deployed; '
                         'model = raw model weights')
    ap.add_argument('--out', default=os.path.join(
        os.path.expanduser('~'), 'Downloads',
        'SQE August_2026 AllIndices Portfolio.csv'))
    args = ap.parse_args()

    port, stamp = load_portfolio(args.data, args.universe)
    lots, deployed = size(port, args.capital)

    # Weight is a percentage (10 = 10%). On the 'executed' basis it is each
    # position's rupee cost over the capital actually deployed, so the column
    # mirrors the real basket rather than the pre-rounding model. The raw inputs
    # carry their own rounding drift, so renormalise and then apportion in
    # 0.01% units by largest remainder: the column sums to exactly 100 without
    # nudging any one holding off its true share.
    raw = ([c / deployed for _, c in lots] if args.basis == 'executed'
           else [h['weight'] for h in port])
    scale = 100 * 10 ** DP
    total = sum(raw)
    exact = [v / total * scale for v in raw]
    units = [int(e) for e in exact]
    order = sorted(range(len(port)), key=lambda i: exact[i] - units[i],
                   reverse=True)
    for i in order[:scale - sum(units)]:
        units[i] += 1

    rows = []
    for h, u, (qty, cost) in zip(port, units, lots):
        sym = h['clean_symbol']
        name, segment = META.get(sym, (sym, h.get('sector', '')))
        rows.append([sym, u / 10 ** DP, segment,
                     rationale(h, name, segment, qty, cost, cost / deployed)])
    rows.sort(key=lambda r: r[1], reverse=True)

    header = ['NSE Ticker', 'Weight', 'Segment (optional)', 'Rationale (optional)']
    if args.out.lower().endswith('.xlsx'):
        # Same four columns as the CSV -- smallcase takes the CSV, this is for
        # reading and circulating. Weight stays numeric so it still sums.
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = 'Portfolio'
        for j, c in enumerate(header, 1):
            cell = ws.cell(1, j, c)
            cell.fill = PatternFill('solid', fgColor='1F3864')
            cell.font = Font(color='FFFFFF', bold=True)
        for i, (sym, wt, segment, why) in enumerate(rows, 2):
            ws.cell(i, 1, sym)
            ws.cell(i, 2, wt).number_format = f'0.{"0" * DP}'
            ws.cell(i, 3, segment)
            ws.cell(i, 4, why).alignment = Alignment(wrap_text=True, vertical='top')
        r = len(rows) + 2
        ws.cell(r, 1, 'TOTAL').font = Font(bold=True)
        ws.cell(r, 2, sum(x[1] for x in rows)).number_format = f'0.{"0" * DP}'
        ws.cell(r, 2).font = Font(bold=True)
        for j, w_ in enumerate([14, 10, 32, 120], 1):
            ws.column_dimensions[get_column_letter(j)].width = w_
        ws.freeze_panes = 'A2'
        wb.save(args.out)
    else:
        with open(args.out, 'w', newline='', encoding='utf-8-sig') as f:
            w = csv.writer(f)
            w.writerow(header)
            for sym, wt, segment, why in rows:
                w.writerow([sym, f'{wt:.{DP}f}', segment, why])

    print(f'[smallcase] {len(rows)} holdings, weight sum '
          f'{sum(r[1] for r in rows):.{DP}f}%, basis={args.basis}, '
          f'data as of {stamp}')
    print(f'[smallcase] Rs {args.capital:,.0f} basket deploys Rs {deployed:,.0f} '
          f'({deployed / args.capital - 1:+.1%}) after whole-share rounding')
    thin = [r[0] for r in rows if r[1] < 1.0]
    if thin:
        print(f'[smallcase] below 0.01 (1%) weight, smallcase min: '
              f'{", ".join(thin)}')
    print(f'[smallcase] wrote -> {args.out}')


if __name__ == '__main__':
    main()
