"""
build_hq_backtest_dashboard.py
--------------------------------
Wires the quarterly fundamental-filter + SOM backtest (build_quarterly_backtest.py
in D:\\Live_share\\web_scrap) into the live "High Quality" dashboard tab, per
explicit instruction:
  - Overview KPIs, equity curve, Exec Summary table, and the P&L heatmap
    (driven by monthly_detail) -> BACKTEST numbers.
  - current_portfolio, exec_history, live_performance, and MONTHLY_HOLDINGS
    (the per-month holdings shown when a heatmap cell is clicked) -> UNTOUCHED,
    still the REAL executed-portfolio data (current + all previous months).
  - Adds a "disclaimer" field (rendered as a banner by app.js) noting the
    backtest only covers the locally-scraped Screener.in universe (~3.5k
    stocks, not the full market), so real full-market returns may vary.

Source: D:\\Live_share\\web_scrap\\data\\quarterly_backtest_summary.csv
Output: D:\\Host_portfolio\\hq_data.js  (DASHBOARD_DATA.high_quality + MONTHLY_HOLDINGS.high_quality)
"""
import csv
import json
import os
import sys

import numpy as np
import pandas as pd


def dd_recovery_months(equity):
    peak = equity[0]
    trough_i = None
    max_dd = 0.0
    for i, v in enumerate(equity):
        if v > peak:
            peak = v
        dd = v / peak - 1
        if dd < max_dd:
            max_dd, trough_i = dd, i
    if trough_i is None:
        return 0, False
    trough_val, target = equity[trough_i], equity[: trough_i + 1]
    peak_before = max(target)
    for j in range(trough_i + 1, len(equity)):
        if equity[j] >= peak_before:
            return j - trough_i, False
    return len(equity) - 1 - trough_i, True


def compute_metrics(returns_series, bench_series, rf_annual=0.06):
    """Copied from extract_dashboard_data.py rather than imported: that module
    runs a full data.js regeneration (incl. live price fetches) as a
    side effect of being imported (top-level script, not import-safe), which
    must not fire here."""
    r = pd.Series(returns_series).dropna()
    b = pd.Series(bench_series).dropna()
    if len(r) == 0:
        return {}
    n = len(r)
    equity = (1 + r).cumprod()
    cagr = float(equity.iloc[-1]) ** (12 / n) - 1
    vol = r.std() * np.sqrt(12)
    sharpe = (cagr - rf_annual) / vol if vol > 0 else 0
    mdd = float((equity / equity.cummax() - 1).min())
    recovery, recovery_ongoing = dd_recovery_months(list(equity))
    downside = r[r < 0].std() * np.sqrt(12)
    sortino = (cagr - rf_annual) / downside if downside > 0 else 0
    calmar = cagr / abs(mdd) if mdd != 0 else 0
    wins = (r > 0).sum()
    win_rate = wins / n
    avg_gain = float(r[r > 0].mean()) if wins > 0 else 0
    avg_loss = float(r[r < 0].mean()) if (r < 0).sum() > 0 else 0
    common = r.index.intersection(b.index)
    alpha = (cagr - b.loc[common].mean() * 12) if len(common) > 0 else 0
    return {
        "CAGR": round(cagr * 100, 2), "Volatility": round(vol * 100, 2), "Sharpe": round(sharpe, 2),
        "Sortino": round(sortino, 2), "Calmar": round(calmar, 2), "Max_DD": round(mdd * 100, 2),
        "Recovery_Months": recovery, "Recovery_Ongoing": recovery_ongoing,
        "Win_Rate": round(win_rate * 100, 1), "Avg_Gain": round(avg_gain * 100, 2),
        "Avg_Loss": round(avg_loss * 100, 2), "Alpha": round(alpha * 100, 2),
        "Total_Return": round((float(equity.iloc[-1]) - 1) * 100, 2)
    }


SOM_SUMMARY_XLSX = r"D:\Host_portfolio\SOM_HQ_Quarterly_v2_Summary.xlsx"  # the REAL SOM-run per-month results (v2 = Aug'26 price refresh, saved under a new name since the v1 files were open in Excel)
SOM_CURRENT_XLSX = r"D:\Host_portfolio\SOM_HQ_Quarterly_v2_Current.xlsx"  # the current month's actual book (Symbol/Action/Qty/Prev Qty/Weight/Price)
SOM_MAIN_XLSX = r"D:\Host_portfolio\SOM_HQ_Quarterly_v2.xlsx"             # PM_YYYY-MM per-month sheets
HQ_DATA_JS = r"D:\Host_portfolio\hq_data.js"
DATA_JS = r"D:\Host_portfolio\data.js"        # only for its sector_map

# Where holdings are priced from, in order. hq_quarterly_universe is the folder
# the SOM run itself read, so its closes are the ones the book was sized on; the
# rest cover the bullion sleeve and anything the screen has since dropped.
PRICE_FOLDERS = [
    "hq_quarterly_universe",
    r"D:\Shared folder\portfolio\High_Quality_September",
    "TOTAL_STOCKS",
    "nifty500_host",
]
# The bullion sleeve is stored under neither naming convention.
BULLION_CSV = {"GOLDBEES": "NSE_GOLDBEES, 1D.csv",
               "SILVERBEES": "NSE_SILVERBEES, 1D.csv"}
BENCH_CSV = "NIFTY500_1d.csv"

DISCLAIMER = (
    "Backtest, not a live track record: this High Quality screen (ROCE>18%, "
    "Market Cap>Rs.1500 Cr, P/E<40, YoY/QoQ profit growth>0, Net Cash Flow>0, "
    "Debt/Equity<1) runs 80% stocks / 10% GOLDBEES / 10% SILVERBEES (same fixed "
    "bullion sleeve as SQE-MultiAsset-ProQuant). It is only evaluated over the "
    "~3,500 stocks we have scraped "
    "financials + price history for locally (Screener.in free-tier universe), "
    "not the full listed market. Actual full-market returns may vary from "
    "what's shown here. Current & previous portfolio holdings below remain "
    "the real executed book. Local daily price data currently runs through "
    "~mid-2026; the most recent 1-2 months are not yet shown as realized "
    "returns pending a price-data refresh."
)


def rolling_return(returns, n):
    if len(returns) < n:
        return None
    window = returns[-n:]
    return float(np.prod([1 + r for r in window]) - 1)


def dd_duration_months(equity):
    """Length (months) of the peak-to-trough run for the max drawdown."""
    peak_idx, trough_idx, max_dd, cur_peak_idx = 0, 0, 0.0, 0
    for i, v in enumerate(equity):
        if v > equity[cur_peak_idx]:
            cur_peak_idx = i
        dd = v / equity[cur_peak_idx] - 1
        if dd < max_dd:
            max_dd, peak_idx, trough_idx = dd, cur_peak_idx, i
    return trough_idx - peak_idx


def adv_metrics(returns, bench):
    """Same 25-key shape the (missing) extract_hq.adv_metrics used to
    produce, using standard, verifiable formulas -- rebuilt here rather than
    guessed at, since extract_hq.py is no longer present in this checkout."""
    r = np.array(returns, dtype=float)
    b = np.array(bench, dtype=float)
    n = len(r)
    equity = np.cumprod(1 + r)
    cagr = float(equity[-1] ** (12 / n) - 1) if n else 0.0
    bench_cagr = float(np.prod(1 + b) ** (12 / n) - 1) if n else 0.0
    vol = float(r.std(ddof=1) * np.sqrt(12)) if n > 1 else 0.0
    downside = float(r[r < 0].std(ddof=1) * np.sqrt(12)) if (r < 0).sum() > 1 else 0.0
    rf = 0.06
    sharpe = (cagr - rf) / vol if vol else 0.0
    sortino = (cagr - rf) / downside if downside else 0.0
    running_max = np.maximum.accumulate(equity)
    dd = equity / running_max - 1
    mdd = float(dd.min())
    calmar = cagr / abs(mdd) if mdd else 0.0
    active = r - b
    info_ratio = float(active.mean() / active.std(ddof=1) * np.sqrt(12)) if active.std(ddof=1) else 0.0
    gains, losses = r[r > 0], r[r < 0]
    win_rate = float((r > 0).sum() / n) if n else 0.0
    profit_factor = float(gains.sum() / abs(losses.sum())) if losses.sum() != 0 else (float('inf') if gains.sum() > 0 else 0.0)

    return {
        "CAGR": cagr, "XIRR": cagr, "Abs Return": float(equity[-1] - 1),
        "Alpha vs Bench": cagr - bench_cagr, "Volatility": vol, "Downside Dev": downside,
        "Sharpe": sharpe, "Sortino": sortino, "Calmar": calmar, "Max Drawdown": mdd,
        "DD Duration (M)": float(dd_duration_months(equity)),
        "VaR 95%": float(np.percentile(r, 5)), "VaR 99%": float(np.percentile(r, 1)),
        "CVaR 95%": float(r[r <= np.percentile(r, 5)].mean()) if n else 0.0,
        "CVaR 99%": float(r[r <= np.percentile(r, 1)].mean()) if n else 0.0,
        "Info Ratio": info_ratio, "Win Rate": win_rate, "Profit Factor": profit_factor,
        "Expectancy": float(r.mean()), "Avg Gain": float(gains.mean()) if len(gains) else 0.0,
        "Avg Loss": float(losses.mean()) if len(losses) else 0.0,
        "Rolling 1Y": rolling_return(returns, 12), "Rolling 3Y": rolling_return(returns, 36),
        "Best Month": float(r.max()) if n else 0.0, "Worst Month": float(r.min()) if n else 0.0,
        "Avg Ex-Ante Sharpe": 0.0,
    }


def load_som_months():
    """Read the REAL per-month SOM results (Port Return %, Bench Return %,
    Ex-ante Beta, stock counts) straight from the Sharpe Single Index Model
    run -- NOT the naive equal-weight approximation from
    build_quarterly_backtest.py's own CSV."""
    import openpyxl
    wb = openpyxl.load_workbook(SOM_SUMMARY_XLSX, data_only=True)
    ws = wb["Summary 5Y"]
    hdr_row = next(r for r in range(1, 15) if ws.cell(row=r, column=1).value
                   and "Portfolio" in str(ws.cell(row=r, column=1).value))
    hdr = [str(ws.cell(row=hdr_row, column=c).value or "").replace("\n", " ").strip()
           for c in range(1, ws.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(hdr) if h}

    import re
    month_re = re.compile(r"^\d{4}-\d{2}$")
    months = []
    for r in range(hdr_row + 1, ws.max_row + 1):
        m = ws.cell(row=r, column=idx["Portfolio Month"]).value
        if not m or not month_re.match(str(m).strip()):
            continue  # skips blanks and the trailing "TOTAL" row
        base_cell = ws.cell(row=r, column=idx["Port Return %"]).value
        if base_cell is None:
            # A genuinely blank/errored cell (seen once, unrelated to any
            # known cause) is NOT the same as a real 0% return -- coercing
            # it via `or 0` would silently fabricate a flat month. Skip the
            # month entirely instead of guessing at its return.
            print(f"[qbt-dash] WARNING: {m} has no Port Return % value (blank/error cell) -- excluded, not zeroed.")
            continue
        # Label by the month the return was EARNED. The sheet's "Portfolio
        # Month" is the signal month; its return lands in the month after, and
        # the heatmap plots on Month while hiding the last row as live -- so
        # August's +9.66% was drawn in the July cell and August showed nothing.
        _tm = ws.cell(row=r, column=idx["Trade Month"]).value if "Trade Month" in idx else None
        if _tm and month_re.match(str(_tm).strip()):
            _label = str(_tm).strip()
        else:
            _y, _mo = map(int, str(m).strip().split("-"))
            _label = f"{_y + 1:04d}-01" if _mo == 12 else f"{_y:04d}-{_mo + 1:02d}"
        months.append({
            "Month": _label,
            "Signal_Month": str(m).strip(),
            "Stock_Count": int(ws.cell(row=r, column=idx["Stocks"]).value or 0),
            "Added": int(ws.cell(row=r, column=idx["Added Stocks"]).value or 0),
            "Removed": int(ws.cell(row=r, column=idx["Removed Stocks"]).value or 0),
            "Port_Beta": round(float(ws.cell(row=r, column=idx["Ex-ante Beta"]).value or 0), 4),
            "Base": float(base_cell),
            "Bench": float(ws.cell(row=r, column=idx["Bench Return %"]).value or 0),
            "Ex_Ante_Sharpe": round(float(ws.cell(row=r, column=idx["Ex-ante Sharpe"]).value or 0), 4),
        })

    # Trailing months with Base == exactly 0.0 are SOM's "Future Selection
    # Mode" placeholders (trade month's price data isn't in yet -- basket
    # formed, nothing realized). Keep AT MOST ONE such row at the tail --
    # the frontend heatmap (app.js renderHeatmap) always renders the LAST
    # monthly_detail row as a "live, not yet traded" marker dot rather than
    # a colored return; dropping every placeholder would instead swallow
    # the last genuinely-realized month under that same marker. Any
    # placeholders beyond the first are further future and just dropped.
    first_placeholder = next((i for i, m in enumerate(months) if m["Base"] == 0.0
                               and all(mm["Base"] == 0.0 for mm in months[i:])), len(months))
    months = months[:first_placeholder + 1]

    # If the data turned out fresh enough that even the last month is fully
    # realized (non-zero), there's no natural placeholder left -- synthesize
    # one for the next calendar month (same basket, not yet realized) so the
    # frontend's "last row = still forming" dot doesn't instead swallow a
    # real, meaningful return.
    if months and months[-1]["Base"] != 0.0:
        last = months[-1]
        y, mo = map(int, last["Month"].split("-"))
        next_month = f"{y + 1:04d}-01" if mo == 12 else f"{y:04d}-{mo + 1:02d}"
        months.append({**last, "Month": next_month, "Added": 0, "Removed": 0, "Base": 0.0, "Bench": 0.0})
    return months


def load_sector_map():
    """sector_map out of data.js. Best-effort: an unmapped symbol renders as
    'Other', which is what the frozen portfolio already showed for most rows."""
    try:
        txt = open(DATA_JS, encoding="utf-8").read()
        i = txt.index("{", txt.index("DASHBOARD_DATA"))
        return json.JSONDecoder().raw_decode(txt[i:])[0].get("sector_map", {})
    except Exception as e:
        print(f"[hq-dash] WARNING: no sector_map ({e}) -- sectors will read 'Other'.")
        return {}


def load_current_portfolio():
    """Rebuild current_portfolio from the SOM Current workbook.

    This used to be carried forward verbatim from the previous hq_data.js on
    the theory that it was "REAL data" the backtest rebuild must not clobber.
    It was real once, but nothing refreshed it after extract_hq.py went missing,
    so it drifted into a different portfolio entirely: holdings the book no
    longer has, 12 holdings it does have missing (including both bullion legs,
    20% of the book), and 11 rows priced at 0. A row priced 0 is not cosmetic --
    app.js recalcPortInvest() skips it, so its weight silently falls out of the
    investment calculator and reappears as phantom "Cash Left".

    Reading the workbook every run is the only way this stays true. Prices come
    from yfinance so LTP is the live session, falling back to the workbook's own
    Current Price column so a fetch failure degrades to a stale price rather
    than to the 0 that breaks the calculator.
    """
    import openpyxl
    from live_prices import price_symbols

    wb = openpyxl.load_workbook(SOM_CURRENT_XLSX, data_only=True)
    ws = wb[wb.sheetnames[0]]
    sectors = load_sector_map()

    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        if not r or not r[0]:
            continue
        sym, action, chg, tgt, prev, weight, price = (list(r) + [None] * 7)[:7]
        sym = str(sym).replace("_1d_max", "").strip()
        tgt, prev = int(tgt or 0), int(prev or 0)
        # Exits (target 0) are instructions, not holdings -- the nifty tabs drop
        # them the same way, and keeping them would dilute every weight.
        if tgt <= 0:
            continue
        rows.append({"symbol": sym, "action": f"{action} {int(chg or 0)}",
                     "qty": tgt, "prev_qty": prev,
                     "weight": float(weight or 0), "book_price": float(price or 0)})

    # Priced from the settled local CSVs, not yfinance: yfinance hands back a
    # placeholder bar for a session that has not happened yet, and rebuilt just
    # after midnight on 01-09 that made 18 of 22 holdings show a 0.00% day.
    quotes = price_symbols([h["symbol"] for h in rows], PRICE_FOLDERS,
                           aliases=BULLION_CSV)
    for h in rows:
        q = quotes.get(h["symbol"], {})
        if not q.get("ltp"):
            # Never emit 0 -- app.js's recalcPortInvest() skips a holding priced
            # 0, so its weight silently drops out of the qty calculator.
            q = {"ltp": h["book_price"], "prev_close": h["book_price"],
                 "change_pct": 0.0, "mtd_change_pct": 0.0, "date": "N/A"}
            print(f"[hq-dash] WARNING: no local price for {h['symbol']} -- "
                  f"falling back to workbook price {h['book_price']}")
        h.update({
            "clean_symbol": h["symbol"],
            "sector": sectors.get(h["symbol"]) or "Other",
            "status": "Remained",
            "ltp": q["ltp"], "prev_close": q["prev_close"],
            "change_pct": q["change_pct"],
            "mtd_change_pct": q["mtd_change_pct"],
            "value": round(h["qty"] * q["ltp"], 2),
            "date": q["date"],
        })
        h.pop("book_price")

    # The whole point of the rebuild: every holding priced, and the weights that
    # drive the qty calculator actually summing to the book.
    unpriced = [h["symbol"] for h in rows if not h["ltp"]]
    assert not unpriced, f"holdings still unpriced: {unpriced}"
    total_w = sum(h["weight"] for h in rows)
    assert abs(total_w - 1.0) < 0.02, f"weights sum to {total_w:.4f}, not ~1.0"

    deployed = sum(h["value"] for h in rows)
    print(f"[hq-dash] current_portfolio rebuilt from {SOM_CURRENT_XLSX}")
    print(f"[hq-dash]   {len(rows)} holdings | weights sum {total_w:.4f} | "
          f"all priced | book value Rs.{deployed:,.0f}")
    for h in sorted(rows, key=lambda x: -x["weight"]):
        print(f"[hq-dash]   {h['clean_symbol']:<12} qty {h['qty']:>7,} x "
              f"Rs.{h['ltp']:>9,.2f} = Rs.{h['value']:>13,.0f}  "
              f"({h['weight'] * 100:5.2f}%)  {h['action']}")
    return rows


def build_monthly_holdings():
    """MONTHLY_HOLDINGS.high_quality from the workbook's PM_ sheets.

    This used to be carried forward from the previous hq_data.js, so it had
    frozen at 2026-06 -- and its contents came from an older run entirely (only
    16 of 30 names matched the current engine's June book). app.js's
    "Portfolio Changes" tab derives exits by diffing current_portfolio against
    the latest snapshot before the live month, so a stale snapshot meant the
    September book was being compared against JUNE: the six positions actually
    exited in September (RPGLIFE, FORCEMOT, MONARCH, EMCURE, BANCOINDIA,
    SHAKTIPUMP) never showed a sell instruction, while long-gone June names
    showed as freshly sold.

    Keyed by the month the book is HELD (the sheet's portfolio month + 1), which
    is how monthly_detail is labelled and how the heatmap modal looks holdings
    up. Selecting "last month's book" from these keys is the front-end's job,
    and app.js does it by skipping the snapshot identical to the current book.
    """
    import re
    import pandas as pd

    xl = pd.ExcelFile(SOM_MAIN_XLSX)
    out = {}
    for sh in sorted(s for s in xl.sheet_names if re.fullmatch(r"PM_\d{4}-\d{2}", s)):
        y, mo = map(int, sh[3:].split("-"))
        trade = f"{y + 1:04d}-01" if mo == 12 else f"{y:04d}-{mo + 1:02d}"
        d = xl.parse(sh, header=None)
        try:
            hdr = next(i for i in range(len(d))
                       if str(d.iloc[i, 1]).strip() == "Symbol")
        except StopIteration:
            continue
        rows = []
        for _, r in d.iloc[hdr + 1:].iterrows():
            sym = str(r.iloc[1]).strip()
            if sym in ("nan", "", "None"):
                break                      # table ends at the first blank row
            num = lambda i: (float(r.iloc[i])
                             if str(r.iloc[i]) not in ("nan", "", "None") else None)
            w = num(10)
            rows.append({"s": sym,
                         "w": round(w * 100, 2) if w is not None else None,
                         "p": round(num(11), 2) if num(11) is not None else None,
                         "st": str(r.iloc[2]).strip(),
                         "b": round(num(4), 3) if num(4) is not None else None,
                         "e": round(num(7), 3) if num(7) is not None else None})
        if rows:
            out[trade] = sorted(rows, key=lambda x: -(x["w"] or 0))
    return out


def build_live_performance(holdings):
    """Portfolio vs benchmark, today and month-to-date, from the same settled
    closes the holdings were priced on.

    On the first day of a new book both figures are 0: the basket was formed at
    last month's close and has not traded a session yet. That is the correct
    answer, and it is what the previous carried-forward value got wrong.
    """
    from live_prices import aggregate, price_symbols

    port_daily, port_mtd = aggregate(holdings)
    b = price_symbols(["BENCH"], [], aliases={"BENCH": BENCH_CSV})["BENCH"]
    bench_daily, bench_mtd = b["change_pct"], b["mtd_change_pct"]
    if not b["ltp"]:
        print(f"[hq-dash] WARNING: benchmark {BENCH_CSV} unreadable -- "
              f"benchmark returns reported as 0")
    return {
        "portfolio_ret": port_daily, "benchmark_ret": bench_daily,
        "alpha": round(port_daily - bench_daily, 2),
        "portfolio_mtd": port_mtd, "benchmark_mtd": bench_mtd,
        "alpha_mtd": round(port_mtd - bench_mtd, 2),
        "indicator": "up" if port_daily >= 0 else "down",
    }


def main():
    months = load_som_months()
    # The last row may be a still-forming "live" placeholder (Base==0.0, no
    # trade data yet) kept only so the heatmap renders its usual dot marker
    # instead of a colored cell -- exclude it from all statistics, same as
    # the real-portfolio pipeline already does for its own live month.
    is_live = bool(months) and months[-1]["Base"] == 0.0 and len(months) > 1
    realized = months[:-1] if is_live else months

    base = [m["Base"] for m in realized]
    bench = [m["Bench"] for m in realized]

    lm_base = compute_metrics(base, bench)
    lm_bench = compute_metrics(bench, bench)
    lm_bench["Alpha"] = 0.0

    eq_base, eq_bench, cb, cn = [], [], 1.0, 1.0
    for m in months:  # include the live row so the equity chart still plots through it (flat, as-is)
        cb *= (1 + m["Base"]); cn *= (1 + m["Bench"])
        eq_base.append(round(cb, 4)); eq_bench.append(round(cn, 4))

    em_base, em_bench = adv_metrics(base, bench), adv_metrics(bench, bench)
    exec_summary = {k: {"Base": em_base[k], "Bench": em_bench[k]} for k in em_base}

    avg_ex_ante_sr = round(sum(m["Ex_Ante_Sharpe"] for m in realized) / len(realized), 4) if realized else 0.0

    # Preserve the REAL data untouched: current_portfolio, exec_history,
    # stock_correlation, live_performance, and the per-month holdings shown
    # when a heatmap cell is clicked.
    with open(HQ_DATA_JS, encoding="utf-8") as f:
        old_txt = f.read()
    start = old_txt.index("DASHBOARD_DATA.high_quality =")
    start = old_txt.index("{", start)
    depth = 0
    for i in range(start, len(old_txt)):
        if old_txt[i] == "{":
            depth += 1
        elif old_txt[i] == "}":
            depth -= 1
            if depth == 0:
                end = i + 1
                break
    old = json.loads(old_txt[start:end])

    hstart = old_txt.index("MONTHLY_HOLDINGS.high_quality =")
    hstart = old_txt.index("{", hstart)
    depth = 0
    for i in range(hstart, len(old_txt)):
        if old_txt[i] == "{":
            depth += 1
        elif old_txt[i] == "}":
            depth -= 1
            if depth == 0:
                hend = i + 1
                break
    old_holdings = json.loads(old_txt[hstart:hend])

    current = load_current_portfolio()
    monthly_holdings = build_monthly_holdings()
    live_perf = build_live_performance(current)

    universe = {
        "exec_summary": exec_summary,
        "avg_ex_ante_sr": avg_ex_ante_sr,
        "layer_metrics": {"Base": lm_base, "Bench": lm_bench},
        "equity_curves": {"months": [m["Month"] for m in months], "Base": eq_base, "Bench": eq_bench},
        "churning_data": [{"Month": m["Month"], "Stock_Count": m["Stock_Count"],
                            "Base Add": m["Added"], "Base Rem": m["Removed"]} for m in months],
        "heatmaps": {},
        "monthly_detail": months,
        "current_portfolio": current,                     # REAL, re-read from the SOM book every run
        "exec_history": old["exec_history"],              # REAL, untouched
        "stock_correlation": old["stock_correlation"],    # REAL, untouched
        "total_months": len(months),
        # Recomputed, NOT carried forward. Carrying it forward left the tab
        # reporting portfolio_mtd 6.32% on 01-09 -- August's figure, on a
        # September book that had not traded a single session.
        "live_performance": live_perf,
        "disclaimer": DISCLAIMER,
    }

    with open(HQ_DATA_JS, "w", encoding="utf-8") as f:
        f.write("/* High Quality (SOM screener) dashboard data.\n"
                "   Overview/heatmap/equity/exec-summary = fundamental-filter + SOM backtest\n"
                "   (build_quarterly_backtest.py + SOM_hq_quarterly.py), Jun'23-Jul'26.\n"
                "   current_portfolio / exec_history / live_performance / MONTHLY_HOLDINGS\n"
                "   remain the REAL executed-portfolio track record, untouched. */\n")
        f.write("DASHBOARD_DATA.high_quality = " + json.dumps(universe, separators=(",", ":"), ensure_ascii=False) + ";\n")
        f.write("if (typeof MONTHLY_HOLDINGS !== 'undefined') MONTHLY_HOLDINGS.high_quality = "
                + json.dumps(monthly_holdings, separators=(",", ":"), ensure_ascii=False) + ";\n")

    print(f"[hq-dash] {len(months)} backtest months -> {HQ_DATA_JS}")
    print(f"[hq-dash] Base CAGR={lm_base['CAGR']}% Sharpe={lm_base['Sharpe']} MaxDD={lm_base['Max_DD']}%")
    print(f"[hq-dash] current_portfolio ({len(universe['current_portfolio'])} stocks) and "
          f"MONTHLY_HOLDINGS ({len(monthly_holdings)} months, "
          f"{min(monthly_holdings)}..{max(monthly_holdings)}) rebuilt from the SOM book.")


if __name__ == "__main__":
    main()
