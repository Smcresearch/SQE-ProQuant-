"""
build_portfolio_xlsx.py
-----------------------
Excel of a current book, laid out like the site's "Current Holdings" table:

    #  Symbol  Sector  Weight  LTP  Today %  Qty  Amount

Holdings and prices come from the live site payload, so the sheet always
matches what the dashboard is showing. Sizing is the same rule app.js uses
(recalcPortInvest): qty = max(1, floor(capital * weight / ltp)).

That 1-share floor is why the deployed total overshoots the basket size: a
high-priced, low-weight name (APARINDS at Rs 17,702 on a 0.94% weight) still
has to buy one share. The overshoot shrinks as the basket grows, so the sheet
reports it explicitly rather than hiding it.

    python build_portfolio_xlsx.py --capital 125000
    python build_portfolio_xlsx.py --universe nifty50 --out x.xlsx
"""
import argparse
import json
import math
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DATA_JS = os.environ.get("SQE_DATA_JS", r"d:/SQE-host/data.js")
LABEL = {"total759": "All Indices", "nifty500": "Nifty 500", "nifty50": "Nifty 50"}

HDR_FILL = PatternFill("solid", fgColor="1F3864")
TOT_FILL = PatternFill("solid", fgColor="DDEBF7")
WHITE = Font(color="FFFFFF", bold=True, size=11)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def load(path, universe):
    s = open(path, encoding="utf-8").read()
    i = s.index("{", s.index("DASHBOARD_DATA"))
    d = json.JSONDecoder().raw_decode(s[i:])[0]
    return d[universe]["current_portfolio"], d.get("last_update", "")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--data", default=DATA_JS)
    ap.add_argument("--universe", default="total759")
    ap.add_argument("--capital", type=float, default=125000)
    ap.add_argument("--out", default=None)
    args = ap.parse_args()

    port, stamp = load(args.data, args.universe)
    port = sorted(port, key=lambda h: -h["weight"])
    name = LABEL.get(args.universe, args.universe)
    out = args.out or f"D:/SQE September_2026 {name.replace(' ', '')} Portfolio.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Current Holdings"

    ws["A1"] = f"SQE {name} - Current Holdings (September 2026)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = (f"Basket Rs {args.capital:,.0f}  |  prices as of {stamp}  |  "
                f"qty = max(1, floor(capital x weight / LTP)), the same rule the site uses")
    ws["A2"].font = Font(size=9, color="595959")

    cols = ["#", "Symbol", "Sector", "Weight", "LTP", "Today %", "Qty", "Amount"]
    for j, c in enumerate(cols, 1):
        cell = ws.cell(4, j, c)
        cell.fill, cell.font, cell.border = HDR_FILL, WHITE, BORDER
        cell.alignment = Alignment(horizontal="center")

    total_amt = 0.0
    for i, h in enumerate(port, 1):
        qty = max(1, math.floor(args.capital * h["weight"] / h["ltp"])) if h["ltp"] else 0
        amt = qty * h["ltp"]
        total_amt += amt
        r = 4 + i
        vals = [i, h["clean_symbol"], h.get("sector") or "Other", h["weight"],
                h["ltp"], (h.get("change_pct") or 0) / 100, qty, amt]
        for j, v in enumerate(vals, 1):
            cell = ws.cell(r, j, v)
            cell.border = BORDER
        ws.cell(r, 4).number_format = "0.00%"
        ws.cell(r, 5).number_format = '"Rs "#,##0.00'
        ws.cell(r, 6).number_format = "+0.00%;-0.00%"
        ws.cell(r, 7).number_format = "#,##0"
        ws.cell(r, 8).number_format = '"Rs "#,##0'

    r = 5 + len(port)
    ws.cell(r, 3, "Total Invested").font = Font(bold=True)
    ws.cell(r, 4, sum(h["weight"] for h in port)).number_format = "0.00%"
    ws.cell(r, 8, total_amt).number_format = '"Rs "#,##0'
    for j in range(1, 9):
        ws.cell(r, j).fill, ws.cell(r, j).border = TOT_FILL, BORDER
        ws.cell(r, j).font = Font(bold=True)

    diff = args.capital - total_amt
    lbl = "Cash Left" if diff >= 0 else "Extra Needed (min 1 share each)"
    ws.cell(r + 1, 3, lbl).font = Font(bold=True)
    c = ws.cell(r + 1, 8, diff)
    c.number_format = '"Rs "#,##0;"-Rs "#,##0'
    c.font = Font(bold=True, color="C00000" if diff < 0 else "006100")

    for j, w in enumerate([5, 14, 34, 10, 13, 10, 8, 13], 1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A5"

    wb.save(out)
    over = total_amt / args.capital - 1
    print(f"[xlsx] {name}: {len(port)} holdings, weights {sum(h['weight'] for h in port):.4f}, "
          f"prices as of {stamp}")
    print(f"[xlsx] Rs {args.capital:,.0f} basket deploys Rs {total_amt:,.0f} ({over:+.1%})")
    print(f"[xlsx] wrote -> {out}")


if __name__ == "__main__":
    main()
